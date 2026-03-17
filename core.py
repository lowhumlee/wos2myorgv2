"""
core.py — WoS MUV Affiliation Ingestion Tool
Core processing engine shared by CLI and Streamlit GUI.
Medical University of Varna · Research Information Systems
"""

from __future__ import annotations

import csv
import difflib
import io
import json
import logging
import os
import re
import sqlite3
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

# ── Import the new initial-aware matcher ──────────────────────────────────────
from initial_matching import InitialAwareMatcher, group_wos_authors

logger = logging.getLogger("wos_muv.core")

# ─── Default Configuration ────────────────────────────────────────────────────

DEFAULT_CONFIG: dict = {
    "muv_affiliation_patterns": [
        "medical university varna",
        "med univ varna",
        "mu varna",
        "medical university of varna",
        "муварна",
        "медицинскиуниверситетварна",
    ],
    "fuzzy_threshold": 0.85,
    "interactive_mode": True,
    "allow_multi_org": True,
    "output_dir": "output",
    "db_path": "staging.db",
    # New settings for initial-expansion matching
    "initial_expansion_enabled": True,
    "initial_expansion_auto_confirm": False,
    "sibling_grouping_enabled": True,
}


def load_config(path: str = "config.json") -> dict:
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            cfg = json.load(f)
        return {**DEFAULT_CONFIG, **cfg}
    return DEFAULT_CONFIG.copy()


# ─── Name Normalization ───────────────────────────────────────────────────────

def strip_diacritics(text: str) -> str:
    if not text: return ""
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )

def normalize_name(name: str) -> str:
    """Basic normalization: lower, no diacritics, only alphanumeric."""
    if not name: return ""
    s = strip_diacritics(name).lower()
    s = re.sub(r'[^a-z0-9\s,]', '', s)
    return " ".join(s.split())

def get_initials_key(name: str) -> str:
    """
    Converts 'Lazarov, Nikola R.' or 'Lazarov, N.R.' into 'lazarov n r'.
    Used for strict matching when full names are unavailable.
    """
    norm = normalize_name(name)
    if ',' not in norm:
        return norm
    surname, given = norm.split(',', 1)
    initials = " ".join([part[0] for part in given.split() if part])
    return f"{surname.strip()} {initials}".strip()

def name_similarity(a: str, b: str) -> float:
    """Fuzzy similarity between two strings."""
    return difflib.SequenceMatcher(None, a, b).ratio()


# ─── Data Parsing ─────────────────────────────────────────────────────────────

def build_person_index(csv_content: str) -> tuple[List[Dict], int, set]:
    """
    Parses ResearcherAndDocument.csv.
    Returns: (list of unique persons, max PersonID found, existing_pairs).

    existing_pairs is a set of (PersonID, DocumentID) tuples already in MyOrg —
    used by batch_process to skip re-uploading documents already linked to a person.

    Each person dict contains:
      "OrganizationID"  - first org ID seen (kept for backward compat)
      "OrganizationIDs" - deduplicated list of ALL org IDs for this person
    """
    persons = {}
    max_pid = 0
    existing_pairs: set[tuple[str, str]] = set()
    f = io.StringIO(csv_content.strip())
    reader = csv.DictReader(f)
    for row in reader:
        pid_str = row.get("PersonID")
        if not pid_str:
            continue

        try:
            pid_int = int(pid_str)
            if pid_int > max_pid:
                max_pid = pid_int
        except ValueError:
            pass

        oid    = row.get("OrganizationID", "").strip()
        doc_id = row.get("DocumentID", "").strip()

        # Always record this (person, document) pair as already existing
        if doc_id:
            existing_pairs.add((pid_str, doc_id))

        if pid_str in persons:
            # Person already recorded — just accumulate extra org IDs
            if oid and oid not in persons[pid_str]["OrganizationIDs"]:
                persons[pid_str]["OrganizationIDs"].append(oid)
            continue

        first_name = row.get("FirstName", "")
        last_name  = row.get("LastName",  "")
        full_name  = f"{last_name}, {first_name}"
        norm       = normalize_name(full_name)

        norm_last  = re.sub(r"[^a-z0-9\s]", "", strip_diacritics(last_name.lower().strip()))
        norm_first = re.sub(r"[^a-z0-9\s]", "", strip_diacritics(first_name.lower().strip()))

        is_init  = all(len(p) == 1 for p in norm_first.split())
        initials = "".join([p[0] for p in norm_first.split() if p])

        persons[pid_str] = {
            "PersonID":        pid_str,
            "AuthorFullName":  full_name,
            "FullName":        full_name,
            "NormName":        norm,
            "Surname":         norm_last,
            "GivenName":       norm_first,
            "Initials":        initials,
            "IsInitialsOnly":  is_init,
            "InitialsKey":     get_initials_key(full_name),
            "OrganizationID":  oid,
            "OrganizationIDs": [oid] if oid else [],
        }
    return list(persons.values()), max_pid, existing_pairs


def build_researcher_dataframe(csv_content: str):
    """
    Builds a pandas DataFrame from ResearcherAndDocument.csv content.
    Used by InitialAwareMatcher. Returns None if pandas is unavailable.
    """
    try:
        import pandas as pd
        f = io.StringIO(csv_content.strip())
        return pd.read_csv(f)
    except Exception:
        return None


def parse_org_hierarchy(csv_content: str) -> List[Dict]:
    """
    Returns list of organization dictionaries.
    Expected by app.py: [{'OrganizationID': '...', 'OrganizationName': '...', 'ParentOrgaID': '...'}]
    """
    orgs = []
    f = io.StringIO(csv_content.strip())
    reader = csv.DictReader(f)
    for row in reader:
        if row.get("OrganizationID"):
            orgs.append({
                "OrganizationID": row.get("OrganizationID"),
                "OrganizationName": row.get("OrganizationName", ""),
                "ParentOrgaID": row.get("ParentOrgaID", "")
            })
    return orgs

def parse_wos_csv(csv_content: str) -> List[Dict]:
    """Parses WoS Export."""
    f = io.StringIO(csv_content.strip())
    sample = csv_content[:2000]
    dialect = 'excel-tab' if '\t' in sample else 'excel'
    reader = csv.DictReader(f, dialect=dialect)
    return [row for row in reader if row.get("UT")]


# ─── Extraction Logic ─────────────────────────────────────────────────────────

def _is_muv_affiliation(affil_norm: str, patterns: list[str]) -> bool:
    """
    Returns True if the normalised affiliation string matches ANY configured
    MUV pattern OR any of the hard-coded fallback variants that cover the
    official full institution name and reversed word-order forms used in WoS.

    Hard-coded fallbacks (in addition to config patterns):
      - "med univ prof dr paraskev stoyanov varna"  (official full name)
      - "varna med univ"                            (reversed word order)
      - "paraskev stoyanov varna"                   (partial official name)
    """
    FALLBACK_PATTERNS = [
        "med univ prof dr paraskev stoyanov varna",
        "med univ prof dr paraskev stoyanov",   # without city (dept follows directly)
        "varna med univ",
        "paraskev stoyanov varna",
        "paraskev stoyanov",                    # minimal form without city
    ]
    all_patterns = list(patterns) + FALLBACK_PATTERNS
    return any(p in affil_norm for p in all_patterns)


def extract_muv_author_pairs(wos_records: List[Dict], cfg: dict) -> List[Dict]:
    """
    Extracts one consolidated row per (author, UT) for every WoS author who
    has at least one MUV-matching affiliation block in the C1 field.

    Key behaviours
    --------------
    * An author who appears in N different MUV C1 blocks (e.g. listed under
      both "Dept Physiol" and "Vasc Biol Res Grp") produces exactly ONE row
      with all MUV affiliation strings merged into muv_affils — avoiding the
      duplicate-row / inflated-count bug.
    * Matching uses the configured patterns PLUS hard-coded fallbacks for the
      official full institution name and reversed word-order forms that WoS
      sometimes exports.
    """
    patterns = [p.lower() for p in cfg.get("muv_affiliation_patterns", []) if p.strip()]

    # key: (author_name, ut) → list of unique MUV affiliation strings
    merged: dict[tuple[str, str], list[str]] = {}

    for rec in wos_records:
        ut = (rec.get("UT") or "").strip()
        c1 = rec.get("C1", "")
        if not c1 or not ut:
            continue

        blocks = re.findall(r'\[(.*?)\]\s*([^\[]+)', c1)
        for authors_str, affil_str in blocks:
            affil_norm = normalize_name(affil_str)
            if not _is_muv_affiliation(affil_norm, patterns):
                continue
            authors = [a.strip() for a in authors_str.split(';') if a.strip()]
            affil_clean = affil_str.strip()
            for auth in authors:
                key = (auth, ut)
                if key not in merged:
                    merged[key] = []
                if affil_clean not in merged[key]:
                    merged[key].append(affil_clean)

    extracted = []
    for (auth, ut), muv_affils in merged.items():
        extracted.append({
            "author_full": auth,
            "AuthorName":  auth,
            "RawAffil":    muv_affils[0],   # primary affil for legacy fields
            "muv_affils":  muv_affils,
            "ut":          ut,
            "UT":          ut,
        })
    return extracted


# ─── Matching Engine ──────────────────────────────────────────────────────────

def match_person(author_name: str, person_index: List[Dict], threshold: float | dict,
                 initial_matcher: InitialAwareMatcher | None = None) -> tuple[str, list]:
    """
    3-tier author matching:
      1. Exact string match   → returns ("exact", [...])
      2. Initial expansion    → WoS "Lazarov, N." matches master "Lazarov, Nikolay R."
                                returns ("fuzzy", [...]) with match_type tag "initial_expansion"
      3. Fuzzy similarity     → SequenceMatcher above threshold
                                returns ("fuzzy", [...])
      4. No match             → returns ("new", [])

    The InitialAwareMatcher (initial_matcher) handles tiers 1–3 when available.
    Falls back to the original logic if it was not built (e.g. pandas missing).
    """
    if isinstance(threshold, dict):
        threshold = float(threshold.get("fuzzy_threshold", 0.85))

    # ── Tier 1–3: use InitialAwareMatcher when available ─────────────────────
    if initial_matcher is not None:
        result = initial_matcher.match(author_name)

        if result.kind == "exact":
            # Convert to the existing tuple format: (score, person_dict, score)
            c = result.candidates[0]
            # Find the full person dict from person_index
            p = next((p for p in person_index if p["PersonID"] == c.person_id), None)
            if p:
                return "exact", [(1.0, p, 1.0)]

        if result.kind in ("initial_expansion", "fuzzy"):
            candidates = []
            for c in result.candidates:
                p = next((p for p in person_index if p["PersonID"] == c.person_id), None)
                if p:
                    # Tag the person dict with match sub-type so the UI can label it
                    tagged_p = {**p, "_match_subtype": c.match_type}
                    candidates.append((c.score, tagged_p, c.score))
            if candidates:
                return "fuzzy", candidates

        return "new", []

    # ── Fallback: original logic (no pandas / InitialAwareMatcher not built) ──
    norm_auth = normalize_name(author_name)
    if ',' not in norm_auth:
        for p in person_index:
            if p["NormName"] == norm_auth:
                return "exact", [(1.0, p, 1.0)]
        return "new", []

    auth_sur, auth_given = norm_auth.split(',', 1)
    auth_sur   = auth_sur.strip()
    auth_given = auth_given.strip()

    auth_initials = "".join([p[0] for p in auth_given.split() if p])
    auth_is_init  = all(len(p) == 1 for p in auth_given.split())

    candidates = []
    for p in person_index:
        if p["NormName"] == norm_auth:
            return "exact", [(1.0, p, 1.0)]
        if p["Surname"] != auth_sur:
            continue
        if not p["Initials"] or not auth_initials or p["Initials"][0] != auth_initials[0]:
            continue
        if auth_initials.startswith(p["Initials"]) or p["Initials"].startswith(auth_initials):
            score = 0.95 if auth_initials == p["Initials"] else 0.90
            candidates.append((score, p, score))
            continue
        if not auth_is_init and not p["IsInitialsOnly"]:
            score = name_similarity(norm_auth, p["NormName"])
            if score >= threshold:
                candidates.append((score, p, score))

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return "fuzzy", candidates

    return "new", []


def group_new_authors(new_records: List[Dict]) -> List[Dict]:
    """
    Groups new authors that are variants of each other before insertion.
    Uses the new group_wos_authors() for consistent sibling detection,
    then falls back to the original logic for any edge cases.
    """
    if not new_records:
        return []

    names = [r.get("author_full", "") for r in new_records]

    # Use the new sibling grouper from initial_matching.py
    sibling_groups = group_wos_authors(names)

    # Build a lookup: name → canonical name for this group
    # The canonical name is the longest (most complete) name in each group
    name_to_canonical: dict[str, str] = {}
    for group_members in sibling_groups.values():
        # Pick the member with the most name parts as canonical
        canonical = max(group_members, key=lambda n: len(normalize_name(n)))
        for member in group_members:
            name_to_canonical[member] = canonical

    processed = []
    for rec in new_records:
        name = rec.get("author_full", "")
        rec["GroupedName"] = name_to_canonical.get(name, name)
        processed.append(rec)

    return processed


# ─── Batch Processing ─────────────────────────────────────────────────────────

def batch_process(muv_pairs: List[Dict], person_index: List[Dict],
                  orgs: List[Dict] | Dict, cfg: dict,
                  start_pid: int = 0,
                  researcher_csv_content: str = "",
                  existing_pairs: set | None = None):
    """
    Processes extracted pairs against the person index.

    Parameters
    ----------
    researcher_csv_content : str
        Raw CSV content of ResearcherAndDocument.csv.
        Used to build the InitialAwareMatcher.
    existing_pairs : set of (PersonID, DocumentID) tuples
        Already-uploaded combinations from ResearcherAndDocument.csv.
        Any WoS pair whose (matched_PersonID, UT) is in this set is skipped
        as already present in MyOrg.

    Returns a dict with 'confirmed', 'needs_review', 'new_persons',
    and 'already_uploaded' (list of skipped pairs).
    """
    confirmed        = []
    needs_review     = []
    already_uploaded = []
    new_persons_staged = {}

    existing_pairs = existing_pairs or set()
    pid_counter = start_pid
    threshold   = float(cfg.get("fuzzy_threshold", 0.85))

    # ── Build InitialAwareMatcher once for the whole batch ───────────────────
    initial_matcher: InitialAwareMatcher | None = None
    if cfg.get("initial_expansion_enabled", True) and researcher_csv_content:
        researcher_df = build_researcher_dataframe(researcher_csv_content)
        if researcher_df is not None:
            initial_matcher = InitialAwareMatcher(researcher_df, fuzzy_threshold=threshold)
            logger.info("InitialAwareMatcher built successfully.")
        else:
            logger.warning("pandas unavailable — falling back to original matching logic.")

    # ── Group pairs by author name ────────────────────────────────────────────
    author_groups: dict[str, list] = defaultdict(list)
    for pair in muv_pairs:
        norm = normalize_name(pair["author_full"])
        author_groups[norm].append(pair)

    # ── First pass: match every unique author name ────────────────────────────
    author_decisions: dict[str, tuple] = {}
    for norm, pairs in author_groups.items():
        author_full = pairs[0]["author_full"]
        match_type, candidates = match_person(
            author_full, person_index, threshold, initial_matcher
        )
        author_decisions[norm] = (match_type, candidates)

    # ── Second pass: group new persons (sibling clustering) ──────────────────
    new_author_records = [
        {"author_full": author_groups[norm][0]["author_full"]}
        for norm, (mt, _) in author_decisions.items()
        if mt == "new"
    ]
    grouped_new      = group_new_authors(new_author_records)
    canonical_lookup = {
        normalize_name(r["author_full"]): r["GroupedName"]
        for r in grouped_new
    }

    # ── Also group fuzzy/initial_expansion candidates as siblings ────────────
    # This ensures "Lazarov, N." and "Lazarov, N. R." appear together in review
    if cfg.get("sibling_grouping_enabled", True):
        unresolved_names = [
            author_groups[norm][0]["author_full"]
            for norm, (mt, _) in author_decisions.items()
            if mt in ("fuzzy", "new")
        ]
        sibling_map = group_wos_authors(unresolved_names)
        # Invert: name → canonical group key (shortest key = most compact label)
        name_to_sibling_group: dict[str, str] = {
            name: gkey
            for gkey, members in sibling_map.items()
            for name in members
        }
    else:
        name_to_sibling_group = {}

    # ── Final pass: build output rows ─────────────────────────────────────────
    for norm, pairs in author_groups.items():
        match_type, candidates = author_decisions[norm]
        author_full = pairs[0]["author_full"]

        if match_type == "exact":
            p = candidates[0][1]
            for pair in pairs:
                if (p["PersonID"], pair["UT"]) in existing_pairs:
                    already_uploaded.append({
                        **pair,
                        "PersonID":       p["PersonID"],
                        "AuthorFullName": p["AuthorFullName"],
                        "OrganizationID": p.get("OrganizationID", ""),
                        "match_type":     "exact",
                        "Reason":         "Already in MyOrg",
                    })
                else:
                    confirmed.append({
                        **pair,
                        "match_type":    "exact",
                        "PersonID":      p["PersonID"],
                        "AuthorFullName": p["AuthorFullName"],
                        "OrganizationID": p.get("OrganizationID", ""),
                    })

        elif match_type == "fuzzy":
            # Determine the display label for the match sub-type
            subtype = candidates[0][1].get("_match_subtype", "fuzzy") if candidates else "fuzzy"
            sibling_group = name_to_sibling_group.get(author_full, author_full)

            # Carry the matched person's org IDs as the suggested default selection
            top_person = candidates[0][1] if candidates else {}
            suggested_org_ids = top_person.get("OrganizationIDs") or (
                [top_person["OrganizationID"]] if top_person.get("OrganizationID") else []
            )

            top_score = candidates[0][0] if candidates else 0.0
            top_pid   = top_person.get("PersonID", "")

            for pair in pairs:
                # High-confidence single-candidate match that is already in existing_pairs
                # → treat as probable duplicate rather than sending to open review
                if (len(candidates) == 1
                        and top_score >= 0.90
                        and top_pid
                        and (top_pid, pair["UT"]) in existing_pairs):
                    already_uploaded.append({
                        **pair,
                        "PersonID":       top_pid,
                        "AuthorFullName": top_person.get("AuthorFullName", author_full),
                        "OrganizationID": top_person.get("OrganizationID", ""),
                        "match_type":     "probable_duplicate",
                        "match_score":    top_score,
                        "Reason":         f"Probable duplicate — {subtype} match (score {top_score:.2f}) already in MyOrg",
                    })
                else:
                    needs_review.append({
                        **pair,
                        "match_type": subtype,
                        "norm": norm,
                        "candidates": candidates,
                        "suggested_pid":     top_pid,
                        "suggested_name":    top_person.get("AuthorFullName", author_full),
                        "suggested_org_ids": suggested_org_ids,
                        "AuthorFullName":    author_full,
                        "SiblingGroup":      sibling_group,
                    })

        else:  # new
            resolved_name = canonical_lookup.get(norm, author_full)
            canon_norm    = normalize_name(resolved_name)
            sibling_group = name_to_sibling_group.get(author_full, author_full)

            if canon_norm not in new_persons_staged:
                pid = str(pid_counter)
                pid_counter += 1
                new_persons_staged[canon_norm] = {
                    "PersonID": pid,
                    "AuthorFullName": resolved_name,
                }

            pid = new_persons_staged[canon_norm]["PersonID"]

            for pair in pairs:
                needs_review.append({
                    **pair,
                    "match_type": "new",
                    "norm": norm,
                    "PersonID": pid,
                    "AuthorFullName": resolved_name,
                    "suggested_pid": pid,
                    "suggested_name": resolved_name,
                    "SiblingGroup": sibling_group,
                })

    return {
        "confirmed":        confirmed,
        "needs_review":     needs_review,
        "new_persons":      list(new_persons_staged.values()),
        "already_uploaded": already_uploaded,
    }


# ─── Persistence & Helpers ────────────────────────────────────────────────────

class StagingDB:
    def __init__(self, db_path: str):
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self._create_tables()

    def _create_tables(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS persons (
                PersonID TEXT PRIMARY KEY, FullName TEXT, NormName TEXT, IsNew INTEGER, Timestamp TEXT
            );
            CREATE TABLE IF NOT EXISTS affiliations (
                PersonID TEXT, UT TEXT, OrgID TEXT, RawAffil TEXT, SourceFile TEXT, Timestamp TEXT,
                PRIMARY KEY (PersonID, UT, OrgID)
            );
            CREATE TABLE IF NOT EXISTS decisions (
                PersonID TEXT, DecisionType TEXT, Detail TEXT, Timestamp TEXT
            );
            CREATE TABLE IF NOT EXISTS rejected (
                AuthorFullName TEXT, UT TEXT, Reason TEXT, Timestamp TEXT
            );
        """)
        self.conn.commit()

    def upsert_person(self, pid: str, full_name: str, norm: str, is_new: bool = True):
        self.conn.execute(
            "INSERT OR IGNORE INTO persons VALUES (?,?,?,?,?)",
            (pid, full_name, norm, int(is_new), datetime.now().isoformat(timespec="seconds"))
        )
        self.conn.commit()

    def log_decision(self, pid: str, dtype: str, detail: str):
        self.conn.execute(
            "INSERT INTO decisions VALUES (?,?,?,?)",
            (pid, dtype, detail, datetime.now().isoformat(timespec="seconds"))
        )
        self.conn.commit()

    def log_rejected(self, author: str, ut: str, reason: str):
        self.conn.execute(
            "INSERT INTO rejected VALUES (?,?,?,?)",
            (author, ut, reason, datetime.now().isoformat(timespec="seconds"))
        )
        self.conn.commit()


# ─── Export Formatters ────────────────────────────────────────────────────────

def _split_full_name(full_name: str) -> tuple[str, str]:
    """
    Split "Lastname, Firstname" into (first_name, last_name).
    Handles names with or without a comma.
    """
    if "," in full_name:
        last, first = full_name.split(",", 1)
        return first.strip(), last.strip()
    # No comma — treat the whole string as last name, first name empty
    return "", full_name.strip()


def build_upload_csv(affiliations: List[Dict], source_file: str = "manual_entry") -> str:
    """
    Produces a CSV in the exact column order required by WoS My Organization:
      PersonID | FirstName | LastName | OrganizationID | DocumentID
    """
    output = io.StringIO()
    fieldnames = ["PersonID", "FirstName", "LastName", "OrganizationID", "DocumentID"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for aff in affiliations:
        first, last = _split_full_name(aff.get("AuthorFullName", ""))
        writer.writerow({
            "PersonID":       aff.get("PersonID", ""),
            "FirstName":      first,
            "LastName":       last,
            "OrganizationID": aff.get("OrgID", aff.get("OrganizationID", "")),
            "DocumentID":     aff.get("UT", aff.get("DocumentID", "")),
        })
    return output.getvalue()

def build_audit_json(summary: dict, new_persons: list) -> str:
    """Generates the audit JSON structure for export."""
    data = {
        "generated_at": datetime.now().isoformat(),
        "summary": summary,
        "new_persons": new_persons,
    }
    return json.dumps(data, indent=2, ensure_ascii=False)

def build_review_excel(results: List[Dict] | dict, org_hierarchy: List[Dict] | Dict = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    # Handle both list and dict inputs
    if isinstance(results, dict):
        all_items = results.get("confirmed", []) + results.get("needs_review", [])
    else:
        all_items = results

    # Sort by SiblingGroup so variants appear together in the sheet
    all_items = sorted(all_items, key=lambda r: r.get("SiblingGroup", r.get("author_full", "")))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Author Review"

    headers = [
        "Status", "WoS Author", "Detected PersonID", "Existing Name",
        "Match Score", "UT", "Affiliation", "OrganizationID", "APPROVED",
        "SiblingGroup",
    ]
    ws.append(headers)

    # Colour fills: yellow = initial_expansion, blue = fuzzy, green = exact, grey = new
    fills = {
        "initial_expansion": PatternFill("solid", fgColor="FFF3CD"),
        "fuzzy":             PatternFill("solid", fgColor="D1ECF1"),
        "exact":             PatternFill("solid", fgColor="D4EDDA"),
        "new":               PatternFill("solid", fgColor="E2E3E5"),
    }

    for r in all_items:
        m_pid   = r.get("PersonID", "")
        m_name  = r.get("AuthorFullName", "")
        score   = r.get("Score", 1.0 if r.get("match_type") == "exact" else 0.0)
        mtype   = r.get("match_type", "UNKNOWN")

        row_data = [
            mtype.upper(),
            r.get("author_full", r.get("AuthorName", "")),
            m_pid,
            m_name,
            score,
            r.get("UT", r.get("ut", "")),
            "; ".join(r.get("muv_affils", [])) if isinstance(r.get("muv_affils"), list) else r.get("RawAffil", ""),
            r.get("OrganizationID", ""),
            "YES" if mtype == "exact" else "PENDING",
            r.get("SiblingGroup", ""),
        ]
        ws.append(row_data)

        # Apply row colour based on match type
        fill = fills.get(mtype, fills["new"])
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # Bold the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
